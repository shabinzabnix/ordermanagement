from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, delete, and_, or_, text
from database import get_db
from models import (
    Product, Store, HOStockBatch, StoreStockBatch,
    InterStoreTransfer, PurchaseRequest, UploadHistory,
    RCCustomer, AuditLog,
    UploadType, TransferStatus, PurchaseStatus,
)
from auth import get_current_user, require_roles
from pydantic import BaseModel
from typing import Optional
from datetime import datetime, timezone, timedelta
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import json
import math

router = APIRouter()


# ─── Utility ──────────────────────────────────────────────

def _excel(rows, headers, filename):
    """Generic Excel generator. headers = [{"label": "...", "key": "..."}]"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Export"
    hdr_font = Font(bold=True, size=10)
    hdr_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h["label"])
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="left")
    for ri, row in enumerate(rows, 2):
        for ci, h in enumerate(headers, 1):
            ws.cell(row=ri, column=ci, value=row.get(h["key"], ""))
    for col in ws.columns:
        ml = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(ml + 3, 50)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


async def _log(db, user, action, entity_type=None, entity_id=None, details=None):
    db.add(AuditLog(
        user_id=user.get("user_id", 0),
        user_name=user.get("full_name", ""),
        action=action,
        entity_type=entity_type,
        entity_id=str(entity_id) if entity_id else None,
        details=details,
    ))


# ─── Inventory Aging ─────────────────────────────────────

@router.get("/aging/report")
async def aging_report(
    category: str = Query(None),
    location: str = Query(None),
    status: str = Query(None),
    bucket: str = Query(None),
    search: str = Query(None),
    page: int = Query(1, ge=1),
    limit: int = Query(100, ge=1, le=500),
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(require_roles("ADMIN", "HO_STAFF", "DIRECTOR")),
):
    from cache import get_cached, set_cached
    cache_key = f"aging_{location}_{status}_{bucket}_{search}_{page}"
    cached = get_cached(cache_key, ttl=90)
    if cached: return cached
    stores = {s.id: s.store_name for s in (await db.execute(select(Store).where(Store.is_active == True))).scalars().all()}
    # Optimized Pure SQL Aging Report via CTE
    # This offloads calculations (age, bucket, status) to the DB
    base_sql = """
        WITH FirstPurchases AS (
            SELECT product_id, store_id, MIN(purchase_date) as first_date
            FROM purchase_records
            GROUP BY product_id, store_id
        ),
        Sales90d AS (
            SELECT product_name, SUM(quantity) as qty
            FROM sales_records
            WHERE invoice_date >= NOW() - INTERVAL '90 days'
            GROUP BY product_name
        ),
        AllStock AS (
            -- Store Stock
            SELECT 
                s.product_name, s.ho_product_id as product_id, s.batch, s.closing_stock_strips as stock,
                s.cost_value as value, s.store_id, st.store_name as location,
                s.created_at, s.expiry_date
            FROM store_stock_batches s
            JOIN stores st ON s.store_id = st.id
            WHERE s.closing_stock_strips > 0
            UNION ALL
            -- HO Stock
            SELECT 
                h.product_name, h.product_id, h.batch, h.closing_stock as stock,
                h.landing_cost_value as value, NULL as store_id, 'Head Office' as location,
                h.created_at, h.expiry_date
            FROM ho_stock_batches h
            WHERE h.closing_stock > 0
        ),
        AgedStock AS (
            SELECT 
                ask.*,
                EXTRACT(DAY FROM (NOW() - COALESCE(fp.first_date, ask.created_at))) as age_days,
                COALESCE(s90.qty, 0) as sales_90d
            FROM AllStock ask
            LEFT JOIN FirstPurchases fp ON ask.product_id = fp.product_id 
                AND (ask.store_id = fp.store_id OR (ask.store_id IS NULL AND fp.store_id IS NULL))
            LEFT JOIN Sales90d s90 ON ask.product_name = s90.product_name
        ),
        ClassifiedStock AS (
            SELECT *,
                CASE 
                    WHEN age_days <= 30 THEN '0-30'
                    WHEN age_days <= 60 THEN '30-60'
                    WHEN age_days <= 90 THEN '60-90'
                    ELSE '90+'
                END as bucket_name,
                CASE
                    WHEN sales_90d = 0 AND age_days > 60 THEN 'dead'
                    WHEN sales_90d < 5 AND age_days > 30 THEN 'slow'
                    ELSE 'active'
                END as status_name
            FROM AgedStock
        )
        SELECT * FROM ClassifiedStock
    """
    
    # Build dynamic filters
    filters = []
    params = {"limit": limit, "offset": (page - 1) * limit}
    
    if location and location != "all":
        filters.append("location = :location")
        params["location"] = location
    if bucket and bucket != "all":
        filters.append("bucket_name = :bucket")
        params["bucket"] = bucket
    if status and status != "all":
        filters.append("status_name = :status")
        params["status"] = status
    if search:
        filters.append("(product_name ILIKE :search OR product_id ILIKE :search)")
        params["search"] = f"%{search}%"
        
    filtered_sql = base_sql
    if filters:
        filtered_sql += " WHERE " + " AND ".join(filters)
    
    # Run all three queries in parallel to eliminate sequential round trips
    tasks = [
        db.execute(text(f"SELECT COUNT(*) FROM ({filtered_sql}) AS total"), params),
        db.execute(text(filtered_sql + " ORDER BY age_days DESC LIMIT :limit OFFSET :offset"), params),
        db.execute(text(f"""
            SELECT 
                COUNT(*) FILTER (WHERE bucket_name = '0-30') as b0_30,
                COUNT(*) FILTER (WHERE bucket_name = '30-60') as b30_60,
                COUNT(*) FILTER (WHERE bucket_name = '60-90') as b60_90,
                COUNT(*) FILTER (WHERE bucket_name = '90+') as b90plus,
                COUNT(*) FILTER (WHERE status_name = 'dead') as dead_count,
                COUNT(*) FILTER (WHERE status_name = 'slow') as slow_count,
                SUM(value) FILTER (WHERE status_name = 'dead') as dead_value,
                SUM(value) FILTER (WHERE status_name = 'slow') as slow_value
            FROM ({filtered_sql}) AS summary
        """), params)
    ]
    
    results = await asyncio.gather(*tasks)
    
    total = results[0].scalar() or 0
    report_rows = results[1].all()
    s_agg = results[2].first()
    
    items = []
    for r in report_rows:
        items.append({
            "location": r.location, "store_id": r.store_id,
            "product_id": r.product_id, "product_name": r.product_name,
            "batch": r.batch, "stock": float(r.stock or 0.0), 
            "days": int(r.age_days or 0), "value": float(r.value or 0.0),
            "sales_90d": float(r.sales_90d or 0.0),
            "stock_date": r.created_at.isoformat() if r.created_at else None,
            "expiry_date": r.expiry_date.isoformat() if r.expiry_date else None,
            "bucket": r.bucket_name, "status": r.status_name,
        })

    result = {
        "items": items, "total": total, "page": page, "limit": limit,
        "summary": {
            "0-30": s_agg.b0_30 if s_agg else 0,
            "30-60": s_agg.b30_60 if s_agg else 0,
            "60-90": s_agg.b60_90 if s_agg else 0,
            "90+": s_agg.b90plus if s_agg else 0
        },
        "dead_count": s_agg.dead_count if s_agg else 0,
        "slow_count": s_agg.slow_count if s_agg else 0,
        "dead_value": math.floor(float((s_agg.dead_value if s_agg else 0) or 0.0) * 100 + 0.5) / 100.0,
        "slow_value": math.floor(float((s_agg.slow_value if s_agg else 0) or 0.0) * 100 + 0.5) / 100.0,
        "locations": ["Head Office"] + list(stores.values()),
    }
    return set_cached(cache_key, result, ttl=120)


@router.get("/intelligence/summary")
async def intelligence_summary(
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    from cache import get_cached, set_cached
    import asyncio
    is_store_staff = user.get("role") in ("STORE_STAFF", "STORE_MANAGER") and user.get("store_id")
    user_store_id = user.get("store_id") if is_store_staff else None
    cache_key = f"intel_summary_{user_store_id or 'all'}"
    cached = get_cached(cache_key, ttl=120)
    if cached: return cached

    now = datetime.now(timezone.utc)
    d30 = now - timedelta(days=30)
    d60 = now - timedelta(days=60)

    # 1. Parallelize data fetching
    tasks = [
        db.execute(select(Store).where(Store.is_active == True)),
    ]

    # Dead and Slow Moving Stock via SQL Aggregation
    agg_stmt = select(
        func.count(StoreStockBatch.id).filter(StoreStockBatch.created_at < d60).label("dc"),
        func.sum(StoreStockBatch.cost_value).filter(StoreStockBatch.created_at < d60).label("dv"),
        func.count(StoreStockBatch.id).filter(and_(StoreStockBatch.created_at < d30, StoreStockBatch.created_at >= d60)).label("sc"),
        func.sum(StoreStockBatch.cost_value).filter(and_(StoreStockBatch.created_at < d30, StoreStockBatch.created_at >= d60)).label("sv")
    ).where(and_(
        StoreStockBatch.closing_stock_strips > 0,
        or_(StoreStockBatch.sales == 0, StoreStockBatch.sales.is_(None))
    ))

    if user_store_id:
        tasks.append(db.execute(
            select(StoreStockBatch)
            .where(and_(StoreStockBatch.store_id == user_store_id, StoreStockBatch.closing_stock_strips > 0, 
                        or_(StoreStockBatch.sales == 0, StoreStockBatch.sales.is_(None)), 
                        StoreStockBatch.created_at < d30))
            .order_by(StoreStockBatch.created_at.asc()).limit(50)
        ))
        agg_stmt = agg_stmt.where(StoreStockBatch.store_id == user_store_id)
    else:
        tasks.append(db.execute(
            select(StoreStockBatch)
            .where(and_(StoreStockBatch.closing_stock_strips > 0, 
                        or_(StoreStockBatch.sales == 0, StoreStockBatch.sales.is_(None)), 
                        StoreStockBatch.created_at < d30))
            .order_by(StoreStockBatch.created_at.asc()).limit(50)
        ))
    
    tasks.append(db.execute(agg_stmt))

    results = await asyncio.gather(*tasks)
    stores_map = {s.id: s.store_name for s in results[0].scalars().all()}
    
    candidate_items = results[1].scalars().all()
    # Unpack the single aggregate row containing dc, dv, sc, sv
    agg_res = results[2].first()
    dead_count = int(agg_res[0] or 0) if agg_res else 0
    dead_val = float(agg_res[1] or 0.0) if agg_res else 0.0
    slow_count = int(agg_res[2] or 0) if agg_res else 0
    slow_val = float(agg_res[3] or 0.0) if agg_res else 0.0

    dead_items = []
    slow_items = []

    for s in candidate_items:
        days = (now - s.created_at).days if s.created_at else 0
        item = {
            "store": stores_map.get(s.store_id, "Head Office"), "store_id": s.store_id,
            "product_id": s.ho_product_id or s.store_product_id or "",
            "product_name": s.product_name, "batch": s.batch,
            "stock": s.closing_stock_strips, "days": days, "sales": 0,
            "value": float(s.cost_value or 0),
        }
        if days >= 60:
            dead_items.append(item)
        else:
            slow_items.append(item)

    # Transfer recommendations (Simplified: only check if Top Dead Items have sales elsewhere)
    recommendations = []
    if dead_items:
        pids = [d["product_id"] for d in dead_items if d["product_id"]]
        if pids:
            # Check for demand in other stores
            demand_res = await db.execute(
                select(StoreStockBatch.ho_product_id, StoreStockBatch.store_id, StoreStockBatch.sales)
                .where(and_(StoreStockBatch.ho_product_id.in_(pids), StoreStockBatch.sales > 0))
            )
            product_sales = {}
            for r in demand_res.all():
                product_sales.setdefault(r[0], []).append({"store_id": r[1], "sales": r[2]})
            
            for d in dead_items:
                if d["product_id"] in product_sales:
                    for buyer in product_sales[d["product_id"]]:
                        if buyer["store_id"] != d["store_id"]:
                            rec_store_name = stores_map.get(buyer["store_id"], f"Store {buyer['store_id']}")
                            recommendations.append({
                                "product_name": d["product_name"], "product_id": d["product_id"],
                                "from_store": d["store"], "to_store": rec_store_name,
                                "from_store_id": d["store_id"], "to_store_id": buyer["store_id"],
                                "quantity": d["stock"], "reason": f"Dead stock ({d['days']}d) with demand at {rec_store_name}",
                            })
                            break
                if len(recommendations) >= 10: break

    result = {
        "dead_stock": dead_items[:20],
        "slow_moving": slow_items[:20],
        "dead_stock_count": dead_count,
        "dead_stock_value": math.floor(dead_val * 100 + 0.5) / 100.0,
        "slow_moving_count": slow_count,
        "slow_moving_value": math.floor(slow_val * 100 + 0.5) / 100.0,
        "recommendations": recommendations,
    }
    return set_cached(cache_key, result, ttl=120)



# ─── Batch Details ────────────────────────────────────────

@router.get("/stock/check-availability")
async def check_transfer_availability(
    source_store_id: int = Query(...),
    product_id: str = Query(...),
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    """Check available stock at source store for a product."""
    stocks = (await db.execute(
        select(StoreStockBatch.batch, StoreStockBatch.closing_stock_strips, StoreStockBatch.mrp)
        .where(and_(StoreStockBatch.store_id == source_store_id, StoreStockBatch.ho_product_id == product_id))
    )).all()
    total = sum(float(s.closing_stock_strips or 0) for s in stocks)
    batches = [{"batch": s.batch, "stock": float(s.closing_stock_strips or 0), "mrp": s.mrp or 0} for s in stocks if (s.closing_stock_strips or 0) > 0]
    return {"total_available": math.floor(float(total) * 10 + 0.5) / 10.0, "batches": batches}


@router.get("/stock/batch-details/{product_id}")
async def batch_details(
    product_id: str,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    stores = {s.id: s.store_name for s in (await db.execute(select(Store).where(Store.is_active == True))).scalars().all()}
    now = datetime.now(timezone.utc)
    batches = []

    for b in (await db.execute(select(HOStockBatch).where(HOStockBatch.product_id == product_id))).scalars().all():
        days = (now - b.created_at).days if b.created_at else 0
        batches.append({
            "location": "Head Office", "batch": b.batch, "mrp": b.mrp or 0,
            "stock": b.closing_stock, "value": b.landing_cost_value or 0,
            "days": days, "sales": 0,
        })

    for b in (await db.execute(select(StoreStockBatch).where(StoreStockBatch.ho_product_id == product_id))).scalars().all():
        days = (now - b.created_at).days if b.created_at else 0
        batches.append({
            "location": stores.get(b.store_id, f"Store {b.store_id}"),
            "batch": b.batch, "mrp": b.mrp or 0,
            "stock": b.closing_stock_strips, "value": b.cost_value or 0,
            "days": days, "sales": b.sales or 0,
        })

    product = (await db.execute(select(Product).where(Product.product_id == product_id))).scalar_one_or_none()
    return {
        "product_id": product_id,
        "product_name": product.product_name if product else "",
        "batches": batches,
    }


# ─── Excel Exports ────────────────────────────────────────

@router.get("/export/products")
async def export_products(db: AsyncSession = Depends(get_db), user: dict = Depends(get_current_user)):
    products = (await db.execute(select(Product).order_by(Product.product_name))).scalars().all()
    rows = [{"product_id": p.product_id, "product_name": p.product_name, "category": p.category or "",
             "sub_category": p.sub_category or "", "primary_supplier": p.primary_supplier or "",
             "secondary_supplier": p.secondary_supplier or "", "mrp": p.mrp or 0, "ptr": p.ptr or 0,
             "landing_cost": p.landing_cost or 0, "rep": p.rep or ""} for p in products]
    headers = [{"label": "Product ID", "key": "product_id"}, {"label": "Product Name", "key": "product_name"},
               {"label": "Category", "key": "category"}, {"label": "Sub Category", "key": "sub_category"},
               {"label": "Primary Supplier", "key": "primary_supplier"}, {"label": "Secondary Supplier", "key": "secondary_supplier"},
               {"label": "MRP", "key": "mrp"}, {"label": "PTR", "key": "ptr"}, {"label": "Landing Cost", "key": "landing_cost"},
               {"label": "REP", "key": "rep"}]
    await _log(db, user, "Exported products to Excel", "product")
    await db.commit()
    return _excel(rows, headers, "products_export.xlsx")


@router.get("/export/ho-stock")
async def export_ho_stock(db: AsyncSession = Depends(get_db), user: dict = Depends(get_current_user)):
    stocks = (await db.execute(select(HOStockBatch).order_by(HOStockBatch.product_name))).scalars().all()
    rows = [{"product_id": s.product_id, "product_name": s.product_name, "batch": s.batch,
             "mrp": s.mrp or 0, "closing_stock": s.closing_stock or 0, "landing_cost_value": s.landing_cost_value or 0} for s in stocks]
    headers = [{"label": "Product ID", "key": "product_id"}, {"label": "Product Name", "key": "product_name"},
               {"label": "Batch", "key": "batch"}, {"label": "MRP", "key": "mrp"},
               {"label": "Closing Stock", "key": "closing_stock"}, {"label": "Landing Cost Value", "key": "landing_cost_value"}]
    await _log(db, user, "Exported HO stock to Excel", "ho_stock")
    await db.commit()
    return _excel(rows, headers, "ho_stock_export.xlsx")


@router.get("/export/store-stock/{store_id}")
async def export_store_stock(store_id: int, db: AsyncSession = Depends(get_db), user: dict = Depends(get_current_user)):
    stocks = (await db.execute(select(StoreStockBatch).where(StoreStockBatch.store_id == store_id).order_by(StoreStockBatch.product_name))).scalars().all()
    store = (await db.execute(select(Store).where(Store.id == store_id))).scalar_one_or_none()
    sname = store.store_name if store else f"Store_{store_id}"
    rows = [{"ho_id": s.ho_product_id or "", "store_id_col": s.store_product_id or "", "product_name": s.product_name,
             "packing": s.packing, "batch": s.batch, "mrp": s.mrp or 0, "sales": s.sales or 0,
             "closing_stock": s.closing_stock, "strips": round(s.closing_stock_strips, 1), "cost_value": s.cost_value or 0} for s in stocks]
    headers = [{"label": "HO ID", "key": "ho_id"}, {"label": "Store ID", "key": "store_id_col"},
               {"label": "Product Name", "key": "product_name"}, {"label": "Packing", "key": "packing"},
               {"label": "Batch", "key": "batch"}, {"label": "MRP", "key": "mrp"}, {"label": "Sales", "key": "sales"},
               {"label": "Closing Stock", "key": "closing_stock"}, {"label": "Stock (Strips)", "key": "strips"},
               {"label": "Cost Value", "key": "cost_value"}]
    await _log(db, user, f"Exported store stock for {sname}", "store_stock", store_id)
    await db.commit()
    return _excel(rows, headers, f"{sname}_stock_export.xlsx")


@router.get("/export/consolidated")
async def export_consolidated(db: AsyncSession = Depends(get_db), user: dict = Depends(require_roles("ADMIN", "HO_STAFF", "DIRECTOR"))):
    stores_list = (await db.execute(select(Store).where(Store.is_active == True).order_by(Store.store_name))).scalars().all()
    products = (await db.execute(select(Product).order_by(Product.product_name))).scalars().all()
    pids = [p.product_id for p in products]
    ho_map = {r[0]: float(r[1] or 0) for r in (await db.execute(
        select(HOStockBatch.product_id, func.sum(HOStockBatch.closing_stock).label("t"))
        .where(HOStockBatch.product_id.in_(pids)).group_by(HOStockBatch.product_id))).all()}
    ss_raw = (await db.execute(
        select(StoreStockBatch.ho_product_id, StoreStockBatch.store_id, func.sum(StoreStockBatch.closing_stock_strips).label("t"))
        .where(and_(StoreStockBatch.ho_product_id.in_(pids), StoreStockBatch.ho_product_id.isnot(None)))
        .group_by(StoreStockBatch.ho_product_id, StoreStockBatch.store_id))).all()
    ss_map = {}
    for r in ss_raw:
        ss_map.setdefault(r[0], {})[r[1]] = float(r[2] or 0)

    headers = [{"label": "Product ID", "key": "product_id"}, {"label": "Product Name", "key": "product_name"},
               {"label": "Category", "key": "category"}, {"label": "HO Stock", "key": "ho_stock"}]
    for s in stores_list:
        headers.append({"label": s.store_name, "key": f"store_{s.id}"})
    headers.append({"label": "TOTAL", "key": "total"})

    rows = []
    for p in products:
        ho = ho_map.get(p.product_id, 0)
        row = {"product_id": p.product_id, "product_name": p.product_name, "category": p.category or "", "ho_stock": ho}
        total = ho
        for s in stores_list:
            qty = ss_map.get(p.product_id, {}).get(s.id, 0)
            row[f"store_{s.id}"] = math.floor(float(qty) * 10 + 0.5) / 10.0
            total += qty
        row["total"] = math.floor(float(total) * 10 + 0.5) / 10.0
        rows.append(row)

    await _log(db, user, "Exported consolidated stock", "consolidated")
    await db.commit()
    return _excel(rows, headers, "consolidated_stock_export.xlsx")


@router.get("/export/transfers")
async def export_transfers(db: AsyncSession = Depends(get_db), user: dict = Depends(get_current_user)):
    transfers = (await db.execute(select(InterStoreTransfer).order_by(InterStoreTransfer.created_at.desc()))).scalars().all()
    sids = set()
    for t in transfers:
        sids.add(t.requesting_store_id); sids.add(t.source_store_id)
    smap = {s.id: s.store_name for s in (await db.execute(select(Store).where(Store.id.in_(sids)))).scalars().all()} if sids else {}
    rows = [{"id": t.id, "from_store": smap.get(t.source_store_id, ""), "to_store": smap.get(t.requesting_store_id, ""),
             "product_id": t.product_id, "product_name": t.product_name, "batch": t.batch or "",
             "quantity": t.quantity, "status": t.status.value if hasattr(t.status, 'value') else t.status,
             "date": t.created_at.strftime("%Y-%m-%d %H:%M") if t.created_at else ""} for t in transfers]
    headers = [{"label": "Transfer ID", "key": "id"}, {"label": "From Store", "key": "from_store"},
               {"label": "To Store", "key": "to_store"}, {"label": "Product ID", "key": "product_id"},
               {"label": "Product Name", "key": "product_name"}, {"label": "Batch", "key": "batch"},
               {"label": "Quantity", "key": "quantity"}, {"label": "Status", "key": "status"}, {"label": "Date", "key": "date"}]
    await _log(db, user, "Exported transfers", "transfer")
    await db.commit()
    return _excel(rows, headers, "transfers_export.xlsx")


@router.get("/export/purchases")
async def export_purchases(db: AsyncSession = Depends(get_db), user: dict = Depends(get_current_user)):
    purchases = (await db.execute(select(PurchaseRequest).order_by(PurchaseRequest.created_at.desc()))).scalars().all()
    sids = set(p.store_id for p in purchases)
    smap = {s.id: s.store_name for s in (await db.execute(select(Store).where(Store.id.in_(sids)))).scalars().all()} if sids else {}
    rows = [{"id": p.id, "store": smap.get(p.store_id, ""), "product_name": p.product_name,
             "brand": p.brand_name or "", "quantity": p.quantity, "customer": p.customer_name,
             "contact": p.customer_contact, "type": "Registered" if p.is_registered_product else "New",
             "status": p.status.value if hasattr(p.status, 'value') else p.status,
             "date": p.created_at.strftime("%Y-%m-%d %H:%M") if p.created_at else ""} for p in purchases]
    headers = [{"label": "ID", "key": "id"}, {"label": "Store", "key": "store"}, {"label": "Product", "key": "product_name"},
               {"label": "Brand", "key": "brand"}, {"label": "Qty", "key": "quantity"}, {"label": "Customer", "key": "customer"},
               {"label": "Contact", "key": "contact"}, {"label": "Type", "key": "type"},
               {"label": "Status", "key": "status"}, {"label": "Date", "key": "date"}]
    await _log(db, user, "Exported purchases", "purchase")
    await db.commit()
    return _excel(rows, headers, "purchase_requests_export.xlsx")


@router.get("/export/uploads")
async def export_uploads(db: AsyncSession = Depends(get_db), user: dict = Depends(get_current_user)):
    uploads = (await db.execute(select(UploadHistory).order_by(UploadHistory.created_at.desc()))).scalars().all()
    rows = [{"id": u.id, "file_name": u.file_name,
             "upload_type": u.upload_type.value if hasattr(u.upload_type, 'value') else u.upload_type,
             "total": u.total_records, "success": u.success_records, "failed": u.failed_records,
             "date": u.created_at.strftime("%Y-%m-%d %H:%M") if u.created_at else ""} for u in uploads]
    headers = [{"label": "ID", "key": "id"}, {"label": "File Name", "key": "file_name"},
               {"label": "Type", "key": "upload_type"}, {"label": "Total", "key": "total"},
               {"label": "Success", "key": "success"}, {"label": "Failed", "key": "failed"}, {"label": "Date", "key": "date"}]
    return _excel(rows, headers, "upload_history_export.xlsx")


@router.get("/export/aging")
async def export_aging(db: AsyncSession = Depends(get_db), user: dict = Depends(require_roles("ADMIN", "HO_STAFF", "DIRECTOR"))):
    report = await aging_report(db=db, user=user)
    rows = [{"location": i["location"], "product_id": i["product_id"], "product_name": i["product_name"],
             "batch": i["batch"], "stock": i["stock"], "mrp": i["mrp"], "days": i["days"],
             "bucket": i["bucket"], "value": i["value"], "status": i["status"].upper()} for i in report["items"]]
    headers = [{"label": "Location", "key": "location"}, {"label": "Product ID", "key": "product_id"},
               {"label": "Product Name", "key": "product_name"}, {"label": "Batch", "key": "batch"},
               {"label": "Stock", "key": "stock"}, {"label": "MRP", "key": "mrp"},
               {"label": "Days in Inventory", "key": "days"}, {"label": "Aging Bucket", "key": "bucket"},
               {"label": "Value", "key": "value"}, {"label": "Status", "key": "status"}]
    return _excel(rows, headers, "inventory_aging_export.xlsx")


# ─── RC Customers ─────────────────────────────────────────

class CustomerCreate(BaseModel):
    store_id: int
    customer_name: str
    mobile_number: str
    medicine_name: str
    last_purchase_date: Optional[str] = None
    duration_of_medication: int = 0
    days_of_consumption: int = 0


@router.get("/customers")
async def get_customers(
    store_id: int = Query(None),
    search: str = Query(None),
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    query = select(RCCustomer)
    if store_id:
        query = query.where(RCCustomer.store_id == store_id)
    if search:
        query = query.where(
            (RCCustomer.customer_name.ilike(f"%{search}%")) | (RCCustomer.medicine_name.ilike(f"%{search}%"))
        )
    customers = (await db.execute(query.order_by(RCCustomer.customer_name))).scalars().all()
    sids = set(c.store_id for c in customers)
    smap = {s.id: s.store_name for s in (await db.execute(select(Store).where(Store.id.in_(sids)))).scalars().all()} if sids else {}
    return {
        "customers": [{
            "id": c.id, "store_id": c.store_id, "store_name": smap.get(c.store_id, ""),
            "customer_name": c.customer_name, "mobile_number": c.mobile_number,
            "medicine_name": c.medicine_name,
            "last_purchase_date": c.last_purchase_date.isoformat() if c.last_purchase_date else None,
            "duration_of_medication": c.duration_of_medication,
            "days_of_consumption": c.days_of_consumption,
        } for c in customers]
    }


@router.post("/customers")
async def create_customer(
    data: CustomerCreate,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    lpd = None
    if data.last_purchase_date:
        try:
            lpd = datetime.fromisoformat(data.last_purchase_date.replace("Z", "+00:00"))
        except ValueError:
            lpd = datetime.now(timezone.utc)
    customer = RCCustomer(
        store_id=data.store_id, customer_name=data.customer_name,
        mobile_number=data.mobile_number, medicine_name=data.medicine_name,
        last_purchase_date=lpd, duration_of_medication=data.duration_of_medication,
        days_of_consumption=data.days_of_consumption, created_by=user["user_id"],
    )
    db.add(customer)
    await _log(db, user, f"Onboarded RC customer: {data.customer_name}", "customer")
    await db.commit()
    await db.refresh(customer)
    return {"id": customer.id, "message": "Customer onboarded"}


@router.put("/customers/{customer_id}")
async def update_customer(
    customer_id: int, data: CustomerCreate,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    c = (await db.execute(select(RCCustomer).where(RCCustomer.id == customer_id))).scalar_one_or_none()
    if not c:
        raise HTTPException(404, "Customer not found")
    c.customer_name = data.customer_name
    c.mobile_number = data.mobile_number
    c.medicine_name = data.medicine_name
    c.duration_of_medication = data.duration_of_medication
    c.days_of_consumption = data.days_of_consumption
    if data.last_purchase_date:
        try:
            c.last_purchase_date = datetime.fromisoformat(data.last_purchase_date.replace("Z", "+00:00"))
        except ValueError:
            pass
    c.updated_at = datetime.now(timezone.utc)
    await _log(db, user, f"Updated RC customer: {data.customer_name}", "customer", customer_id)
    await db.commit()
    return {"message": "Customer updated"}


@router.delete("/customers/{customer_id}")
async def delete_customer(
    customer_id: int,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    c = (await db.execute(select(RCCustomer).where(RCCustomer.id == customer_id))).scalar_one_or_none()
    if not c:
        raise HTTPException(404, "Customer not found")
    await db.delete(c)
    await _log(db, user, f"Deleted RC customer ID {customer_id}", "customer", customer_id)
    await db.commit()
    return {"message": "Customer deleted"}


@router.get("/customers/refill-reminders")
async def refill_reminders(
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    customers = (await db.execute(select(RCCustomer).where(RCCustomer.days_of_consumption > 0))).scalars().all()
    sids = set(c.store_id for c in customers)
    smap = {s.id: s.store_name for s in (await db.execute(select(Store).where(Store.id.in_(sids)))).scalars().all()} if sids else {}
    now = datetime.now(timezone.utc)
    reminders = []
    for c in customers:
        if not c.last_purchase_date or c.days_of_consumption <= 0:
            continue
        next_refill = c.last_purchase_date + timedelta(days=c.days_of_consumption)
        days_until = (next_refill - now).days
        if days_until <= 7:
            reminders.append({
                "id": c.id, "customer_name": c.customer_name, "mobile_number": c.mobile_number,
                "medicine_name": c.medicine_name, "store_name": smap.get(c.store_id, ""),
                "last_purchase": c.last_purchase_date.isoformat(),
                "next_refill": next_refill.isoformat(), "days_until": days_until,
                "overdue": days_until < 0,
            })
    return {"reminders": sorted(reminders, key=lambda x: x["days_until"])}


# ─── Audit Logs ───────────────────────────────────────────

@router.get("/scorecard")
async def store_scorecard(
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(require_roles("ADMIN", "HO_STAFF", "DIRECTOR")),
):
    """Store Performance Scorecard: ranks stores by turnover, dead stock %, transfer compliance."""
    now = datetime.now(timezone.utc)
    stores = (await db.execute(select(Store).where(Store.is_active == True).order_by(Store.store_name))).scalars().all()
    if not stores:
        return {"stores": [], "network_avg": {}}

    scorecards = []
    for store in stores:
        sid = store.id
        stock_rows = (await db.execute(select(StoreStockBatch).where(StoreStockBatch.store_id == sid))).scalars().all()

        total_stock = sum(float(r.closing_stock_strips or 0) for r in stock_rows)
        total_sales = sum(float(r.sales or 0) for r in stock_rows)
        total_value = sum(float(r.cost_value or 0) for r in stock_rows)
        total_items = len([r for r in stock_rows if r.closing_stock_strips and r.closing_stock_strips > 0])
        dead_items = 0
        slow_items = 0
        total_aging_days = 0
        for r in stock_rows:
            if r.closing_stock_strips and r.closing_stock_strips > 0:
                days = (now - r.created_at).days if r.created_at else 0
                total_aging_days += days
                if (r.sales or 0) == 0 and days > 60:
                    dead_items += 1
                elif (r.sales or 0) == 0 and days > 30:
                    slow_items += 1

        sku_count = len(set((r.ho_product_id or r.store_product_id or r.product_name) for r in stock_rows if r.closing_stock_strips and r.closing_stock_strips > 0))
        turnover = math.floor(float(total_sales / total_stock) * 100 + 0.5) / 100.0 if total_stock > 0 else 0.0
        dead_pct = math.floor(float(dead_items / total_items * 100) * 10 + 0.5) / 10.0 if total_items > 0 else 0.0
        slow_pct = math.floor(float(slow_items / total_items * 100) * 10 + 0.5) / 10.0 if total_items > 0 else 0.0
        avg_aging = math.floor(float(total_aging_days / total_items) + 0.5) if total_items > 0 else 0.0

        # Transfer compliance
        t_all = (await db.execute(
            select(func.count(InterStoreTransfer.id)).where(
                (InterStoreTransfer.source_store_id == sid) | (InterStoreTransfer.requesting_store_id == sid)
            )
        )).scalar() or 0
        t_approved = (await db.execute(
            select(func.count(InterStoreTransfer.id)).where(
                and_(
                    (InterStoreTransfer.source_store_id == sid) | (InterStoreTransfer.requesting_store_id == sid),
                    InterStoreTransfer.status == TransferStatus.APPROVED,
                )
            )
        )).scalar() or 0
        compliance = math.floor(float(t_approved / t_all * 100) * 10 + 0.5) / 10.0 if t_all > 0 else 100.0

        # Purchase request count
        pr_count = (await db.execute(
            select(func.count(PurchaseRequest.id)).where(PurchaseRequest.store_id == sid)
        )).scalar() or 0

        # Composite score: Turnover 40% (cap at 1.0 = 100), Dead Stock 30% (invert), Compliance 30%
        norm_turnover = min(turnover * 100, 100)
        score = math.floor(float(norm_turnover * 0.4 + (100 - dead_pct) * 0.3 + compliance * 0.3) * 10 + 0.5) / 10.0

        scorecards.append({
            "store_id": sid,
            "store_name": store.store_name,
            "store_code": store.store_code,
            "location": store.location or "",
            "sku_count": sku_count,
            "total_stock": math.floor(float(total_stock) * 10 + 0.5) / 10.0,
            "total_sales": math.floor(float(total_sales) * 10 + 0.5) / 10.0,
            "stock_value": math.floor(float(total_value) * 100 + 0.5) / 100.0,
            "turnover_ratio": turnover,
            "dead_stock_pct": dead_pct,
            "slow_moving_pct": slow_pct,
            "dead_items": dead_items,
            "slow_items": slow_items,
            "avg_aging_days": avg_aging,
            "transfer_compliance": compliance,
            "transfers_total": t_all,
            "transfers_approved": t_approved,
            "purchase_requests": pr_count,
            "score": score,
        })

    scorecards.sort(key=lambda x: -x["score"])
    for i, sc in enumerate(scorecards):
        sc["rank"] = i + 1

    # Network averages
    n = len(scorecards) or 1
    network_avg = {
        "avg_turnover": math.floor(float(sum(s["turnover_ratio"] for s in scorecards) / n) * 100 + 0.5) / 100.0,
        "avg_dead_pct": math.floor(float(sum(s["dead_stock_pct"] for s in scorecards) / n) * 10 + 0.5) / 10.0,
        "avg_compliance": math.floor(float(sum(s["transfer_compliance"] for s in scorecards) / n) * 10 + 0.5) / 10.0,
        "avg_score": math.floor(float(sum(s["score"] for s in scorecards) / n) * 10 + 0.5) / 10.0,
        "total_network_value": math.floor(float(sum(s["stock_value"] for s in scorecards)) * 100 + 0.5) / 100.0,
    }

    return {"stores": scorecards, "network_avg": network_avg}


@router.get("/export/scorecard")
async def export_scorecard(db: AsyncSession = Depends(get_db), user: dict = Depends(require_roles("ADMIN", "HO_STAFF", "DIRECTOR"))):
    data = await store_scorecard(db=db, user=user)
    rows = [{
        "rank": s["rank"], "store": s["store_name"], "code": s["store_code"],
        "skus": s["sku_count"], "stock": s["total_stock"], "sales": s["total_sales"],
        "value": s["stock_value"], "turnover": s["turnover_ratio"],
        "dead_pct": s["dead_stock_pct"], "slow_pct": s["slow_moving_pct"],
        "avg_aging": s["avg_aging_days"], "compliance": s["transfer_compliance"],
        "score": s["score"],
    } for s in data["stores"]]
    headers = [
        {"label": "Rank", "key": "rank"}, {"label": "Store", "key": "store"}, {"label": "Code", "key": "code"},
        {"label": "SKUs", "key": "skus"}, {"label": "Stock (Units)", "key": "stock"},
        {"label": "Sales", "key": "sales"}, {"label": "Stock Value", "key": "value"},
        {"label": "Turnover Ratio", "key": "turnover"}, {"label": "Dead Stock %", "key": "dead_pct"},
        {"label": "Slow Moving %", "key": "slow_pct"}, {"label": "Avg Aging (Days)", "key": "avg_aging"},
        {"label": "Transfer Compliance %", "key": "compliance"}, {"label": "Performance Score", "key": "score"},
    ]
    await _log(db, user, "Exported store scorecard", "scorecard")
    await db.commit()
    return _excel(rows, headers, "store_scorecard.xlsx")


@router.get("/dashboard/chart-data")
async def get_chart_data(
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    from cache import get_cached, set_cached
    import asyncio
    is_store = user.get("role") in ("STORE_STAFF", "STORE_MANAGER") and user.get("store_id")
    user_store_id = user.get("store_id") if is_store else None
    cache_key = f"chart_data_{user_store_id or 'all'}"
    cached = get_cached(cache_key, ttl=120)
    if cached: return cached

    now = datetime.now(timezone.utc)
    d30 = now - timedelta(days=30)
    d60 = now - timedelta(days=60)
    d90 = now - timedelta(days=90)

    # Parallelize independent queries
    tasks = [
        db.execute(select(Store).where(Store.is_active == True)),
        # Aging chart (Store Stock)
        db.execute(
            select(
                func.sum(case((StoreStockBatch.created_at >= d30, StoreStockBatch.closing_stock_strips), else_=0)).label("u30"),
                func.sum(case((and_(StoreStockBatch.created_at < d30, StoreStockBatch.created_at >= d60), StoreStockBatch.closing_stock_strips), else_=0)).label("u60"),
                func.sum(case((and_(StoreStockBatch.created_at < d60, StoreStockBatch.created_at >= d90), StoreStockBatch.closing_stock_strips), else_=0)).label("u90"),
                func.sum(case((StoreStockBatch.created_at < d90, StoreStockBatch.closing_stock_strips), else_=0)).label("u90p"),
                func.sum(case((StoreStockBatch.created_at >= d30, StoreStockBatch.cost_value), else_=0)).label("v30"),
                func.sum(case((and_(StoreStockBatch.created_at < d30, StoreStockBatch.created_at >= d60), StoreStockBatch.cost_value), else_=0)).label("v60"),
                func.sum(case((and_(StoreStockBatch.created_at < d60, StoreStockBatch.created_at >= d90), StoreStockBatch.cost_value), else_=0)).label("v90"),
                func.sum(case((StoreStockBatch.created_at < d90, StoreStockBatch.cost_value), else_=0)).label("v90p"),
            ).where(StoreStockBatch.closing_stock_strips > 0 if not user_store_id else and_(StoreStockBatch.closing_stock_strips > 0, StoreStockBatch.store_id == user_store_id))
        ),
        # Stock distribution
        db.execute(
            select(StoreStockBatch.store_id, func.sum(StoreStockBatch.closing_stock_strips).label("t"))
            .where(StoreStockBatch.closing_stock_strips > 0 if not user_store_id else and_(StoreStockBatch.closing_stock_strips > 0, StoreStockBatch.store_id == user_store_id))
            .group_by(StoreStockBatch.store_id)
        ),
        # Top categories
        db.execute(
            select(Product.category, func.count(Product.id).label("cnt"))
            .where(Product.category.isnot(None))
            .group_by(Product.category).order_by(func.count(Product.id).desc()).limit(8)
        ),
        # Transfer status
        db.execute(
            select(InterStoreTransfer.status, func.count(InterStoreTransfer.id).label("cnt"))
            .group_by(InterStoreTransfer.status)
        )
    ]

    # HO Aging (if admin)
    if not user_store_id:
        from sqlalchemy import case
        tasks.append(db.execute(
            select(
                func.sum(case((HOStockBatch.created_at >= d30, HOStockBatch.closing_stock), else_=0)).label("u30"),
                func.sum(case((and_(HOStockBatch.created_at < d30, HOStockBatch.created_at >= d60), HOStockBatch.closing_stock), else_=0)).label("u60"),
                func.sum(case((and_(HOStockBatch.created_at < d60, HOStockBatch.created_at >= d90), HOStockBatch.closing_stock), else_=0)).label("u90"),
                func.sum(case((HOStockBatch.created_at < d90, HOStockBatch.closing_stock), else_=0)).label("u90p"),
                func.sum(case((HOStockBatch.created_at >= d30, HOStockBatch.landing_cost_value), else_=0)).label("v30"),
                func.sum(case((and_(HOStockBatch.created_at < d30, HOStockBatch.created_at >= d60), HOStockBatch.landing_cost_value), else_=0)).label("v60"),
                func.sum(case((and_(HOStockBatch.created_at < d60, HOStockBatch.created_at >= d90), HOStockBatch.landing_cost_value), else_=0)).label("v90"),
                func.sum(case((HOStockBatch.created_at < d90, HOStockBatch.landing_cost_value), else_=0)).label("v90p"),
            ).where(HOStockBatch.closing_stock > 0)
        ))
        tasks.append(db.execute(select(func.sum(HOStockBatch.closing_stock))))

    results = await asyncio.gather(*tasks)
    
    stores_map = {s.id: s.store_name for s in results[0].scalars().all()}
    ss_aging = results[1].one()
    ss_dist = results[2].all()
    cat_q = results[3].all()
    trans_q = results[4].all()

    # Process aging
    buckets = ["0-30", "30-60", "60-90", "90+"]
    u_sums = [float(ss_aging[0] or 0), float(ss_aging[1] or 0), float(ss_aging[2] or 0), float(ss_aging[3] or 0)]
    v_sums = [float(ss_aging[4] or 0), float(ss_aging[5] or 0), float(ss_aging[6] or 0), float(ss_aging[7] or 0)]

    if not user_store_id:
        ho_aging = results[5].one()
        for i in range(4):
            u_sums[i] += float(ho_aging[i] or 0)
            v_sums[i] += float(ho_aging[i+4] or 0)

    aging_chart = [{"bucket": buckets[i], "units": math.floor(float(u_sums[i]) + 0.5), "value": math.floor(float(v_sums[i]) + 0.5)} for i in range(4)]

    # Process distribution
    store_dist = []
    if not user_store_id:
        ho_total = float(results[6].scalar() or 0)
        if ho_total > 0:
            store_dist.append({"name": "Head Office", "value": math.floor(ho_total + 0.5)})
    for r in ss_dist:
        v = float(r[1] or 0)
        if v > 0:
            store_dist.append({"name": stores_map.get(r[0], f"Store {r[0]}"), "value": math.floor(v + 0.5)})

    category_chart = [{"name": r[0] or "Uncategorized", "count": r[1]} for r in cat_q]
    transfer_chart = [{"status": r[0].value if hasattr(r[0], 'value') else r[0], "count": r[1]} for r in trans_q]

    result = {
        "aging_chart": aging_chart,
        "stock_distribution": store_dist,
        "category_chart": category_chart,
        "transfer_chart": transfer_chart,
    }
    return set_cached(cache_key, result, ttl=120)



@router.get("/audit-logs")
async def get_audit_logs(
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=200),
    entity_type: str = Query(None),
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(require_roles("ADMIN")),
):
    query = select(AuditLog)
    if entity_type:
        query = query.where(AuditLog.entity_type == entity_type)
    total = (await db.execute(select(func.count()).select_from(query.subquery()))).scalar()
    logs = (await db.execute(query.order_by(AuditLog.created_at.desc()).offset((page - 1) * limit).limit(limit))).scalars().all()
    return {
        "logs": [{
            "id": l.id, "user_name": l.user_name, "action": l.action,
            "entity_type": l.entity_type, "entity_id": l.entity_id,
            "details": l.details, "created_at": l.created_at.isoformat() if l.created_at else None,
        } for l in logs],
        "total": total, "page": page, "limit": limit,
    }
